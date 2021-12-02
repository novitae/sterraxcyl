import openpyxl
import re
import json
from sterra._pri import outExcept

def getnameinfos(pt):
    # NOM DU FICHIER
    name = pt.split('/')[-1]
    if not '#' in pt:
        # TYPE DE LISTE
        tp = re.findall(r'[a-z]{1,16}\.[A-Za-z]{3}-[0-9]{2}-[0-9]{4};[0-9]{2}-[0-9]{2}', name)[0].split('.')[0]
        # FORMAT
        fm = re.findall(r'(xlsx|csv|json)', name.split('.')[-1])[0]
        # TARGET USERNAME
        tu = '_'.join(re.findall(r'[a-z0-9_]{1,30}_'+tp, pt)[0].split('_')[:-1])
        # PATH
        pt = '/'.join(pt.split('/')[:-1])+'/'

    else:
        splt = name.split('#')
        # TYPE DE LISTE
        tp = splt[0]
        # FORMAT
        fm = re.findall(r'(xlsx|csv|json)', name.split('.')[-1])[0]
        tu = None
        # PATH
        pt = '/'.join(pt.split('/')[:-1])+'/'
    
    return {
        'username': tu,
        'type': tp,
        'path': pt,
        'name': name,
        'format': fm if fm != 'xlsx' else 'excel'
        }

def checkpathformat(pt):
    if re.match(r'[a-z_]{3,48}\.[A-Za-z]{3}-[0-9]{2}-[0-9]{4};[0-9]{2}-[0-9]{2}\.[a-z]{3,4}', getnameinfos(pt)['name']):
        return True
    return False

def excel(tr):
    ExcelFile = openpyxl.load_workbook(tr)
    ExcelSheet = ExcelFile[ExcelFile.sheetnames[0]]
    
    Colonnes = ExcelSheet.max_column
    Rangées = ExcelSheet.max_row
    Retour = []
    
    for r in range(2, Rangées):
        Dic = {}
        for c in range(1, Colonnes):
            Val = ExcelSheet.cell(row=r, column=c).value
            if type(Val) == str:
                if '=HYPERLINK(' in Val:
                    Val = Val.replace('=HYPERLINK(', '').replace(')', '').split(', ')[0].replace('"', '')
            Dic[ExcelSheet.cell(row=1, column=c).value] = Val
        Retour.append(Dic)

    return Retour

def CSV(tr):
    Retour = []

    CSVFile = open(tr, 'r').read()
    CSVTitles = CSVFile.split('\n')[:1][0].split(',')
    CSVLine = CSVFile.split('\n')[1:]
    for r in CSVLine:
        Dic = {}
        c = r.split(',')
        for i, e in enumerate(c):
            Dic[CSVTitles[i]] = e.replace('<breakline>', '\n').replace('<coma>', ',')
        Retour.append(Dic)

    return Retour[:-1]

def Json(tr):
    return json.loads(open(tr, 'r').read())

def verter(entry) -> list:
    '''- entry: can be path of the file to convert to dict, also a dict to convert to the comparaison module format'''
    try:
        ext = entry.split('.')[-1]
        if re.match(r'xlsx', ext):
            return excel(entry)
        elif re.match(r'csv', ext):
            return CSV(entry)
        elif re.match(r'json', ext):
            return Json(entry)
    
    except FileNotFoundError:
        outExcept(e=f'Conversion module didn\'t find the file {entry}!\n    Be sure of your path.')