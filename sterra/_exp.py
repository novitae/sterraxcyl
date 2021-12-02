import openpyxl
import json
import re

from time import time
from datetime import datetime, date
from stringcolor import *

from sterra import _his
from sterra import _pri

class ort:
    def __init__(self, ld, pt, fm, tp=None, tu=None, dt='', cn=False):
        '''- ld: List of ?__a=1 dictionnaries
        - tp: Type of data contained
        - fm: Format in wich the file will be saved
        - tu: Instagram target's username
        - pt: Path where to write the file
        - dt: Data to write on the file infos case for excel
        - cn: Custom name'''

        if not fm in ['csv', 'excel', 'json']:
            raise ValueError('"fm" must be "csv", "excel" or "json".')

        xt = {
            'csv': '.csv',
            'excel': '.xlsx',
            'json': '.json'}

        self.ld = ld
        self.tp = tp
        self.fm = fm
        self.tu = tu
        if not cn:
            try:
                self.fn = pt+self.tu.replace('.', '_')+'_'+self.tp+'.'+date.today().strftime("%b-%d-%Y")+datetime.now().strftime(";%H-%M-%S")+xt[self.fm]
            except AttributeError:
                self.fn = pt+self.tp+'.'+date.today().strftime("%b-%d-%Y")+datetime.now().strftime(";%H-%M-%S")+xt[self.fm]
        else:
            self.fn = f'{pt}{cn}#{int(time())}{xt[self.fm]}'
        self.dt = dt if dt != '' else tp

    def EXcsv(self):
        cycle = 0
        for d in self.ld:
            if cycle == 0:
                open(self.fn, 'a').write(','.join(d.keys())+'\n')
            open(self.fn, 'a').write(','.join([str(d[key]).replace('\n', '<breakline>').replace(',', '<coma>') for key in d.keys()])+'\n')
            cycle += 1

    def MKexcel(self, cn):
        '''Create the excel file
        - cn: Cases name (type == list)'''
        def fileInfos(x):
            ExcelSheet.cell(row = 1, column = x, value = 'S T E R R A')
            ExcelSheet.cell(row = 2, column = x, value = '=HYPERLINK("https://github.com/novitae", "** Made By novitae **")').style='Hyperlink'
            ExcelSheet.cell(row = 4, column = x, value = 'Targeted account :')
            ExcelSheet.cell(row = 5, column = x, value = f'=HYPERLINK("https://www.instagram.com/{self.tu}/", "{self.tu}")').style='Hyperlink'
            ExcelSheet.cell(row = 6, column = x, value = 'Date :')
            ExcelSheet.cell(row = 7, column = x, value = date.today().strftime("%b-%d-%Y")+datetime.now().strftime(" %H:%M:%S"))
            ExcelSheet.cell(row = 9, column = x, value = 'Data :')
            ExcelSheet.cell(row = 10, column = x, value = self.dt)
        
        ExcelFile = openpyxl.Workbook()
        ExcelSheet = ExcelFile[ExcelFile.sheetnames[0]]
        for nx in range(len(cn)):
            ExcelSheet.cell(row=1, column=nx+1, value=cn[nx])
        fileInfos(len(cn)+1)
        ExcelFile.save(self.fn)

    def EXSexcel(self, d, r):
        '''Fills the excel file created before
        - d: Dictionnary
        - r: Row number'''
        r += 2
        ExcelFile = openpyxl.load_workbook(filename = self.fn)
        ExcelSheet = ExcelFile[ExcelFile.sheetnames[0]]
        c = 1
        for v in range(len(d.keys())):
            vl = list(d.values())
            if vl[v] != None:
                if re.findall(r'http[s]{0,1}:\/\/', str(vl[v])) != [] and not ' ' in str(vl[v]):
                    ExcelSheet.cell(row = r, column = c, value = f'=HYPERLINK("{vl[v]}", "{vl[v]}")').style='Hyperlink'
                else:
                    ExcelSheet.cell(row = r, column = c, value = type(vl[v])(vl[v]))
            c += 1
        ExcelFile.save(self.fn)

    def d(self):
        '''Destination'''
        open(self.fn, 'w').close()
        print(f'{_pri.i()}Writing file ...')
        if self.fm == 'csv':
            logo = _pri.CSV()
            ort.EXcsv(self)
        elif self.fm == 'excel':
            logo = _pri.Excel()
            ort.MKexcel(self, list(self.ld[0].keys()))
            for r in range(len(self.ld)):
                ort.EXSexcel(self, self.ld[r], r)
        elif self.fm == 'json':
            logo = _pri.Json()
            open(self.fn, 'w').write(json.dumps(self.ld, indent=4))

        print(f'{logo}File succesfully exported under '+bold(self.fn).underline())
        
        _his.tory(dd=self.fn)
        return self.fn