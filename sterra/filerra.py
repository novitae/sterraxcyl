from json import loads, dumps
from openpyxl import Workbook, load_workbook
from re import findall
from datetime import datetime, date

from sterra._sterrage_ import DAY_STRF, HOUR_STRF
from sterra.exterra import exman

class exporter:
    def __init__(self, _:object, List:list, file_path:str, Format:str, **kwargs:dict) -> None:
        self._ = _

        self.list = List
        self.dict_keys = list(self.list[0].keys())
        self.Format = Format
        self.file_path = file_path
        self.kwargs = kwargs
    
    def __call__(self) -> str:
        getattr(self,self.Format.upper())()
        return # (self.id if self.is_part else add(**{"path":self.fp}))

    def XLSX(self) -> Workbook:
        excelfile = Workbook()
        exclsheet = excelfile.active

        usn = self.kwargs.get("username")

        for ik, dk in enumerate(self.dict_keys):
            exclsheet.cell(row=1, column=ik+1, value=dk)

        default_infos = [
            {"v":'S T E R R A'},
            {"v":'=HYPERLINK("https://github.com/novitae", "** Made By novitae **")',"style":'Hyperlink'},
            {"v":'Date :'},
            {"v":f"{date.today().strftime(DAY_STRF)}:{datetime.now().strftime(HOUR_STRF)}"}
            ]
        if self.kwargs.get("target"):
            default_infos.append({"v":"Containing :"})
            default_infos.append({"v":self.kwargs.get("target")})
        if usn:
            default_infos.append({"v":'Targeted account :'})
            default_infos.append({"v":f'=HYPERLINK("https://www.instagram.com/{usn}/", "{usn}")',"style":'Hyperlink'})

        for infnum, inf in enumerate(default_infos):
            exclsheet.cell(row=infnum+1, column=len(self.dict_keys)+1, value=inf["v"]).style=(inf["style"] if "style" in list(inf.keys()) else "Normal")

        for ligne, dico in enumerate(self.list):
            for colonne, val in enumerate(list(dico.values())):
                if val or type(val) is bool:
                    exclsheet.cell(row=ligne+2, column=colonne+1, value=val).style=("Hyperlink" if findall(r'http[s]{0,1}:\/\/', str(val)) and " " not in val else "Normal")

        excelfile.save(self.file_path)

    def CSV(self) -> open:
        with open(self.file_path,"w") as w:
            w.write(','.join(self.dict_keys)+'\n'+"\n".join([",".join([str(v).replace('\n','</b>').replace(',','</c>') for v in d.values()]) for d in self.list]))

    def JSON(self) -> open:
        open(self.file_path,"w").write(dumps(self.list,indent=4))

class reader:
    def __init__(self, _:object, file_path:str) -> None:
        self._ = _
        self.file_path = file_path
        self.f:str = exman(_)._decompose_path(self.file_path)["format"]

    def __call__(self) -> list:
        return getattr(self,self.f.upper())()

    def XLSX(self) -> list:
        excelfile = load_workbook(self.file_path)
        exclsheet = excelfile.active

        maxRow = exclsheet.max_row #! MaxRow prends la longueur de la présentation; si la liste ne contient les données d'un seul username, le maxrow sera toujours de 9
        maxCol = exclsheet.max_column

        rtr = []
        for r in range(2,maxRow):
            dico = {}
            for c in range(1, maxCol):
                val = exclsheet.cell(row=r, column=c).value
                val = val if type(val) != str else val.replace('=HYPERLINK(', '').replace(')', '').split(', ')[0].replace('"', '')
                # Check le type inutile mais on sait jamais si certains chargent des fichiers avec des igid en int
                    
                dico[exclsheet.cell(row=1, column=c).value] = val
            
            if list(dico.values()) != [None]*len(dico): #! Pour éviter l'erreur de maxRow ligne 81
                rtr.append(dico)
        return rtr

    def CSV(self) -> list:
        with open(self.file_path,"r") as r:
            lines = r.read().split("\n")
            keys = lines[0].split(",")
            rtr = [{keys[n]:v.replace('</b>','\n').replace('</c>',',') for n, v in enumerate(preDict.split(","))} for preDict in lines[1:]]
        return rtr

    def JSON(self) -> list:
        return loads(open(self.file_path,"r").read())

# Remplir la case "containg" quand on fait une conversion
