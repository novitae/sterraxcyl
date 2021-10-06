from datetime import datetime, date
import openpyxl, csv, os
from stringcolor import *

def MakeExcel(inst): #CRÉATION DU FICHIER EXCEL
    def fileInfos(x):
        ExcelSheet.cell(row = 1, column = x, value = 'S T E R R A X C Y L')
        ExcelSheet.cell(row = 2, column = x, value = '=HYPERLINK("{}", "{}")'.format('https://github.com/novitae', '** Made By novitae **')).style='Hyperlink'
        ExcelSheet.cell(row = 4, column = x, value = 'Targeted account :')
        ExcelSheet.cell(row = 5, column = x, value = '=HYPERLINK("https://www.instagram.com/{}/", "{}")'.format(inst['username'], inst['username'])).style='Hyperlink'
        ExcelSheet.cell(row = 6, column = x, value = 'List :')
        ExcelSheet.cell(row = 7, column = x, value = inst['target'])
        ExcelSheet.cell(row = 9, column = x, value = 'Date :')
        ExcelSheet.cell(row = 10, column = x, value = date.today().strftime("%b-%d-%Y")+datetime.now().strftime(" %H:%M"))
    
    ExcelFile = openpyxl.Workbook()
    ExcelSheet = ExcelFile[ExcelFile.sheetnames[0]]
    ExcelSheet.cell(row = 1, column = 1, value = 'ID')
    ExcelSheet.cell(row = 1, column = 2, value = 'Username')
    ExcelSheet.cell(row = 1, column = 3, value = 'FullName')
    ExcelSheet.cell(row = 1, column = 4, value = 'Page')
    ExcelSheet.cell(row = 1, column = 5, value = 'Biography')
    ExcelSheet.cell(row = 1, column = 6, value = 'IsPrivate')
    ExcelSheet.cell(row = 1, column = 7, value = 'Followers')
    ExcelSheet.cell(row = 1, column = 8, value = 'Following')
    ExcelSheet.cell(row = 1, column = 9, value = 'Posts')
    ExcelSheet.cell(row = 1, column = 10, value = 'External Link')
    ExcelSheet.cell(row = 1, column = 11, value = 'IsBusiness')
    ExcelSheet.cell(row = 1, column = 12, value = 'IsProfessional')
    ExcelSheet.cell(row = 1, column = 13, value = 'IsVerified')
    if inst['a_allinfos']:
        ExcelSheet.cell(row = 1, column = 14, value = 'Business Adress')
        ExcelSheet.cell(row = 1, column = 15, value = 'Business Category')
        ExcelSheet.cell(row = 1, column = 16, value = 'Business Contact Method')
        ExcelSheet.cell(row = 1, column = 17, value = 'Business Email')
        ExcelSheet.cell(row = 1, column = 18, value = 'Business Phone Number')
        ExcelSheet.cell(row = 1, column = 19, value = 'Connected Facebook Page')
        ExcelSheet.cell(row = 1, column = 20, value = 'Mutual Followed By')
        ExcelSheet.cell(row = 1, column = 21, value = 'Has Effects')
        ExcelSheet.cell(row = 1, column = 22, value = 'Has Channel')
        ExcelSheet.cell(row = 1, column = 23, value = 'Has Clips')
        ExcelSheet.cell(row = 1, column = 24, value = 'Has Guide')
        ExcelSheet.cell(row = 1, column = 25, value = 'Hide Like and View Count')
        ExcelSheet.cell(row = 1, column = 26, value = 'Has joined Recently')
        fileInfos(27)
    else:
        fileInfos(14)
    ExcelFile.save(inst['path']+inst['username'].replace('.', '_')+'_'+inst['target']+'.xlsx')

def ExportExcel(inst, InfoList, rowN): #EXPORT DES DONNÉES VERS LE FICHIER EXCEL
    fileName = inst['path']+inst['username'].replace('.', '_')+'_'+inst['target']+'.xlsx'
    ExcelFile = openpyxl.load_workbook(filename = fileName)
    ExcelSheet = ExcelFile[ExcelFile.sheetnames[0]]
    ExcelSheet.cell(row = rowN, column = 1, value = int(InfoList['Id']))
    ExcelSheet.cell(row = rowN, column = 2, value = InfoList['Username'])
    ExcelSheet.cell(row = rowN, column = 3, value = InfoList['FullName'])
    ExcelSheet.cell(row = rowN, column = 4, value = '=HYPERLINK("{}", "{}")'.format('https://www.instagram.com/{}/'.format(InfoList['Username']), "Account")).style='Hyperlink'
    ExcelSheet.cell(row = rowN, column = 5, value = InfoList['Biography'])
    ExcelSheet.cell(row = rowN, column = 6, value = InfoList['IsPrivate'])
    ExcelSheet.cell(row = rowN, column = 7, value = int(InfoList['Followers']))
    ExcelSheet.cell(row = rowN, column = 8, value = int(InfoList['Follows']))
    ExcelSheet.cell(row = rowN, column = 9, value = int(InfoList['Posts']))
    if InfoList['Link'] == None:
        pass
    else:
        ExcelSheet.cell(row = rowN, column = 10, value = '=HYPERLINK("{}", "{}")'.format(InfoList['Link'], InfoList['Link'])).style='Hyperlink'
    ExcelSheet.cell(row = rowN, column = 11, value = InfoList['IsBusiness'])
    ExcelSheet.cell(row = rowN, column = 12, value = InfoList['IsProfessional'])
    ExcelSheet.cell(row = rowN, column = 13, value = InfoList['IsVerified'])
    if inst['a_allinfos']:
        ExcelSheet.cell(row = rowN, column = 14, value = InfoList['BAdress'])
        ExcelSheet.cell(row = rowN, column = 15, value = InfoList['BCategory'])
        ExcelSheet.cell(row = rowN, column = 16, value = InfoList['BContactMethod'])
        ExcelSheet.cell(row = rowN, column = 17, value = InfoList['BEmail'])
        ExcelSheet.cell(row = rowN, column = 18, value = InfoList['BPhone'])
        if InfoList['CFacebook'] == None:
            pass
        else:
            ExcelSheet.cell(row = rowN, column = 19, value = '=HYPERLINK("{}", "{}")'.format(InfoList['CFacebook'], InfoList['CFacebook'])).style='Hyperlink'
        ExcelSheet.cell(row = rowN, column = 20, value = InfoList['MutualFollowedBy'])
        ExcelSheet.cell(row = rowN, column = 21, value = InfoList['HasAREffect'])
        ExcelSheet.cell(row = rowN, column = 22, value = InfoList['HasChannel'])
        ExcelSheet.cell(row = rowN, column = 23, value = InfoList['HasClips'])
        ExcelSheet.cell(row = rowN, column = 24, value = InfoList['HasGuide'])
        ExcelSheet.cell(row = rowN, column = 25, value = InfoList['HideLAndV'])
        ExcelSheet.cell(row = rowN, column = 26, value = InfoList['JoinedRecently'])
    
    ExcelFile.save(fileName)        

def ExportCSV(inst, InfoList): #EXPORTATION DES DONNÉES VERS LE FICHIER CSV
    path = inst['path']+inst['username'].replace('.', '_')+'_'+inst['target']+'.csv'
    if inst['a_allinfos']:
        fieldnames = ['ID', 'Username', 'FullName', 'Page', 'Biography', 'IsPrivate', 'Followers', 'Following', 'Posts', 'External Link', 'IsBusiness', 'IsProfessional', 'IsVerified', 'Business Adress', 'Business Category', 'Business Contact Method', 'Business Email', 'Business Phone Number', 'Connected Facebook Page', 'Mutual Followed By', 'Has Effects', 'Has Channel', 'Has Clips', 'Has Guide', 'Hide Like and View Count', 'Has joined Recently']
    else:
        fieldnames = ['ID', 'Username', 'FullName', 'Page', 'Biography', 'IsPrivate', 'Followers', 'Following', 'Posts', 'External Link', 'IsBusiness', 'IsProfessional', 'IsVerified']
    writer = csv.DictWriter(open(path, 'a', newline=''), fieldnames = fieldnames)

    if inst['a_allinfos']:
        writer.writerow({'ID': InfoList['Id'], 'Username': InfoList['Username'], 'FullName': InfoList['FullName'], 'Page': "https://www.instagram.com/{}/".format(InfoList['Username']), 'Biography': InfoList['Biography'], 'IsPrivate': InfoList['IsPrivate'], 'Followers': InfoList['Followers'], 'Following': InfoList['Follows'], 'Posts': InfoList['Posts'], 'External Link': InfoList['Link'], 'IsBusiness': InfoList['IsBusiness'], 'IsProfessional': InfoList['IsProfessional'], 'IsVerified': InfoList['IsVerified'], 'Business Adress': InfoList['BAdress'], 'Business Category': InfoList['BCategory'], 'Business Contact Method': InfoList['BContactMethod'], 'Business Email': InfoList['BEmail'], 'Business Phone Number': InfoList['BPhone'], 'Connected Facebook Page': InfoList['CFacebook'], 'Mutual Followed By': InfoList['MutualFollowedBy'], 'Has Effects': InfoList['HasAREffect'], 'Has Channel': InfoList['HasChannel'], 'Has Clips': InfoList['HasClips'], 'Has Guide': InfoList['HasGuide'], 'Hide Like and View Count': InfoList['HideLAndV'], 'Has joined Recently': InfoList['JoinedRecently']})
    else:
        writer.writerow({'ID': InfoList['Id'], 'Username': InfoList['Username'], 'FullName': InfoList['FullName'], 'Page': "https://www.instagram.com/{}/".format(InfoList['Username']), 'Biography': InfoList['Biography'], 'IsPrivate': InfoList['IsPrivate'], 'Followers': InfoList['Followers'], 'Following': InfoList['Follows'], 'Posts': InfoList['Posts'], 'External Link': InfoList['Link'], 'IsBusiness': InfoList['IsBusiness'], 'IsProfessional': InfoList['IsProfessional'], 'IsVerified': InfoList['IsVerified']})

def ConvertInfos(inst, I, rowN): #CONVERSION DU JSON
    try:
        I = I["graphql"]["user"]
    except KeyError:
        print('error not supposed to happen, report this in the issue section of sterraxcyl on github')
        print(I)
        exit()

    InfoList = {
        'Biography': I["biography"],
        'Link': I["external_url"],
        'Followers': I["edge_followed_by"]["count"],
        'Follows': I["edge_follow"]["count"],
        'FullName': I["full_name"],
        'Id': I["id"],
        'IsBusiness': I["is_business_account"],
        'IsProfessional': I["is_professional_account"],
        'IsPrivate': I["is_private"],
        'IsVerified': I["is_verified"],
        'Username': I["username"],
        'Posts': I["edge_owner_to_timeline_media"]["count"]}
    if inst['a_allinfos']:
        InfoList = {
            'Biography': I["biography"],
            'Link': I["external_url"],
            'Followers': I["edge_followed_by"]["count"],
            'Follows': I["edge_follow"]["count"],
            'FullName': I["full_name"],
            'Id': I["id"],
            'IsBusiness': I["is_business_account"],
            'IsProfessional': I["is_professional_account"],
            'IsPrivate': I["is_private"],
            'IsVerified': I["is_verified"],
            'Username': I["username"],
            'Posts': I["edge_owner_to_timeline_media"]["count"],
            'BAdress': I["business_address_json"],
            'BCategory': I["business_category_name"],
            'BContactMethod': I["business_contact_method"],
            'BEmail': I["business_email"],
            'BPhone': I["business_phone_number"],
            'CFacebook': I["connected_fb_page"],
            'MutualFollowedBy': I["edge_mutual_followed_by"]["count"],
            'HasAREffect': I["has_ar_effects"],
            'HasChannel': I["has_channel"],
            'HasClips': I["has_clips"],
            'HasGuide': I["has_guides"],
            'HideLAndV': I["hide_like_and_view_counts"],
            'JoinedRecently': I["is_joined_recently"]}

    ExportCSV(inst, InfoList) if inst['a_csv'] else ExportExcel(inst, InfoList, rowN)

if __name__ == '__main__':
    print('This file is a module for Sterraxcyl, you can\'t open it by itself.')